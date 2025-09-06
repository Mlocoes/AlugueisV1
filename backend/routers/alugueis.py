# Correção: importar e definir router corretamente

from config import get_db
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query, Form
from sqlalchemy.orm import Session, joinedload
import pandas as pd
from typing import Optional
from datetime import datetime
from models_final import Imovel, Proprietario, AluguelSimples, Usuario
from sqlalchemy import asc, desc, func
from .auth import verify_token
import calendar
# Assuming CalculoService is in this path
from services.calculo_service import CalculoService

router = APIRouter(prefix="/api/alugueis", tags=["alugueis"])

@router.post("/importar/", response_model=dict)
async def importar_alugueis(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Importa aluguéis de um arquivo Excel em formato matriz.
    Estrutura: cada linha = imóvel, 1ª coluna = nome do imóvel, 2ª coluna ignorada,
    da 3ª até penúltima coluna = nome do proprietário (célula = valor do aluguel líquido),
    última coluna = taxa de administração total do imóvel.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="O arquivo deve ser do tipo Excel (.xlsx ou .xls)")
    try:
        conteudo = await file.read()
        from io import BytesIO
        import openpyxl
        erros = []
        detalhes_abas = []
        total_processados = 0
        total_criados = 0

        # Carregar workbook com openpyxl
        wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            celula_a1 = ws['A1'].value
            # Carregar DataFrame normalmente (header=0), processar dados a partir da linha 2 (A2)
            df = pd.read_excel(BytesIO(conteudo), sheet_name=nome_aba, header=0)
            colunas = df.columns.tolist()
            if len(colunas) < 4:
                erros.append(f"Aba '{nome_aba}': Planilha deve ter pelo menos 4 colunas: imóvel, ignorada, pelo menos um proprietário e taxa de administração.")
                continue
            # Segunda coluna (B) = ignorar, nomes dos proprietários = C até penúltima, última = taxa
            nomes_proprietarios = colunas[2:-1]
            nome_col_taxa = colunas[-1]
            # Los nombres de los propietarios están como strings, no datetime
            nomes_proprietarios = [str(prop) for prop in nomes_proprietarios if not isinstance(prop, datetime)]

            lidos = 0
            importados = 0
            ignorados = 0
            # Obter data de referência da célula A1
            try:
                if isinstance(celula_a1, datetime):
                    ano = celula_a1.year
                    mes = celula_a1.month
                else:
                    data_ref = pd.to_datetime(str(celula_a1), errors='coerce')
                    if pd.isna(data_ref):
                        ano = datetime.now().year
                        mes = datetime.now().month
                    else:
                        ano = data_ref.year
                        mes = data_ref.month
                print(f"🟢 Aba '{nome_aba}': Data lida da célula A1 = {celula_a1} | ano = {ano}, mes = {mes}")
            except Exception as e:
                ano = datetime.now().year
                mes = datetime.now().month
                erros.append(f"Aba '{nome_aba}': Erro ao obter data de referência da célula A1: {str(e)}")
                print(f"🔴 Aba '{nome_aba}': Erro ao obter data de referência da célula A1: {str(e)}")

            for index, row in df.iterrows():
                # El nombre del inmueble está en la primera columna de datos, no el header
                nome_imovel = str(row.iloc[0]).strip()
                
                # Verificar si es un nombre válido de inmueble
                if pd.isna(row.iloc[0]) or nome_imovel == 'nan' or nome_imovel == '':
                    continue
                    
                imovel = db.query(Imovel).filter(Imovel.nome == nome_imovel).first()
                if not imovel:
                    erros.append(f"Aba '{nome_aba}' Linha {index + 2}: Imóvel '{nome_imovel}' não encontrado.")
                    lidos += 1
                    ignorados += 1
                    continue

                taxa_adm_total = row[nome_col_taxa]
                for nome_prop in nomes_proprietarios:
                    valor_liquido = row[nome_prop]
                    if pd.isna(valor_liquido) or valor_liquido == 0:
                        continue
                    proprietario = db.query(Proprietario).filter(Proprietario.nome == nome_prop).first()
                    if not proprietario:
                        erros.append(f"Aba '{nome_aba}' Linha {index + 2}: Proprietário '{nome_prop}' não encontrado.")
                        ignorados += 1
                        continue
                    dados_aluguel = {
                        "imovel_id": imovel.id,
                        "proprietario_id": proprietario.id,
                        "mes": int(mes),
                        "ano": int(ano),
                        "valor_liquido_proprietario": float(valor_liquido) if not pd.isna(valor_liquido) else 0,
                        "taxa_administracao_total": float(taxa_adm_total) if not pd.isna(taxa_adm_total) else 0,
                    }
                    try:
                        # Tenta inserir, se já existir faz update (upsert manual)
                        existente = db.query(AluguelSimples).filter_by(
                            imovel_id=imovel.id,
                            proprietario_id=proprietario.id,
                            mes=mes,
                            ano=ano
                        ).first()
                        if existente:
                            # Atualiza apenas colunas existentes no modelo
                            existente.valor_liquido_proprietario = float(valor_liquido) if not pd.isna(valor_liquido) else existente.valor_liquido_proprietario
                            existente.taxa_administracao_total = float(taxa_adm_total) if not pd.isna(taxa_adm_total) else existente.taxa_administracao_total
                            # Outros campos do modelo podem ser ajustados se necessário
                            db.commit()
                            importados += 1
                        else:
                            novo_aluguel = AluguelSimples(**dados_aluguel)
                            db.add(novo_aluguel)
                            db.commit()
                            db.refresh(novo_aluguel)
                            importados += 1
                    except Exception as e:
                        db.rollback()
                        ignorados += 1
                        erros.append(f"Aba '{nome_aba}' Linha {index + 2}: Erro ao criar/atualizar aluguel para '{nome_prop}': {str(e)}")
                lidos += 1

            detalhes_abas.append({
                "aba": nome_aba,
                "lidos": lidos,
                "importados": importados,
                "ignorados": ignorados,
                "mensagem": f"Aba '{nome_aba}': {lidos} registros lidos, {importados} importados, {ignorados} ignorados (duplicados ou erro)."
            })
            total_processados += lidos
            total_criados += importados

        total_ignorados = sum(aba.get("ignorados", 0) for aba in detalhes_abas)
        mensagem_final = (
            f"Importação concluída com sucesso.\n"
            f"Total de registros lidos: {total_processados}\n"
            f"Total importados: {total_criados}\n"
            f"Total ignorados: {total_ignorados}\n"
            f"Veja detalhes por aba abaixo."
        )

        return {
            "total_lidos": total_processados,
            "total_importados": total_criados,
            "total_ignorados": total_ignorados,
            "erros": len(erros),
            "detalhe_errores": erros[:10],
            "abas": detalhes_abas,
            "mensagem": mensagem_final
        }
    except Exception as e:
        db.rollback()
        print(f"Erro na importação de aluguéis: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar o arquivo: {str(e)}")


def formatar_periodo_label(ano: int, mes: int) -> str:
    """Formatar período em formato legível"""
    try:
        nome_mes = calendar.month_name[mes] if mes and 1 <= mes <= 12 else str(mes)
        return f"{nome_mes} {ano}"
    except:
        return f"{mes}/{ano}"

@router.get("/listar")
async def listar_alugueis(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(2000, ge=1, le=10000, description="Número máximo de registros a retornar (padrão: 2000, máx: 10000)"),
    ano: Optional[int] = Query(None, ge=2020, le=2030, description="Filtrar por ano"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Filtrar por mês"),
    imovel_id: Optional[int] = Query(None, description="Filtrar por ID do imóvel"),
    proprietario_id: Optional[int] = Query(None, description="Filtrar por ID do proprietário"),
    ordem: str = Query("desc", description="Ordem: 'asc' ou 'desc'"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Listar aluguéis com filtros e paginação"""
    try:
        query = db.query(AluguelSimples).options(
            joinedload(AluguelSimples.imovel),
            joinedload(AluguelSimples.proprietario)
        )
        # Aplicar filtros
        if ano:
            query = query.filter(AluguelSimples.ano == ano)
        if mes:
            query = query.filter(AluguelSimples.mes == mes)
        if imovel_id:
            query = query.filter(AluguelSimples.imovel_id == imovel_id)
        if proprietario_id:
            query = query.filter(AluguelSimples.proprietario_id == proprietario_id)
        # Aplicar ordem
        if ordem.lower() == "asc":
            query = query.order_by(asc(AluguelSimples.ano), asc(AluguelSimples.mes), asc(AluguelSimples.imovel_id))
        else:
            query = query.order_by(desc(AluguelSimples.ano), desc(AluguelSimples.mes), asc(AluguelSimples.imovel_id))
        # Aplicar paginação
        alugueis = query.offset(skip).limit(limit).all()
        # Devolver também nomes relacionados
        data = [aluguel.to_dict() for aluguel in alugueis]
        return {"success": True, "data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao listar aluguéis: {str(e)}")

@router.get("/obter/{aluguel_id}")
async def obter_aluguel(aluguel_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Obter um aluguel específico por ID"""
    try:
        aluguel = db.query(AluguelSimples).filter(AluguelSimples.id == aluguel_id).first()
        if not aluguel:
            raise HTTPException(status_code=404, detail="Aluguel não encontrado")
        data = {
            **aluguel.to_dict(),
            'nome_imovel': aluguel.imovel.nome if aluguel.imovel else None,
            'nome_proprietario': aluguel.proprietario.nome if aluguel.proprietario else None
        }
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao obter aluguel: {str(e)}")

@router.post("/criar")
async def criar_aluguel(
    ano: int = Form(...),
    mes: int = Form(...),
    imovel_id: int = Form(...),
    proprietario_id: int = Form(...),
    valor: float = Form(...),
    descricao: str = Form(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Criar um novo aluguel"""
    try:
        # Verificar se já existe um aluguel para o mesmo imóvel/proprietário/ano/mês
        aluguel_existente = db.query(AluguelSimples).filter(
            AluguelSimples.ano == ano,
            AluguelSimples.mes == mes,
            AluguelSimples.imovel_id == imovel_id,
            AluguelSimples.proprietario_id == proprietario_id
        ).first()
        
        if aluguel_existente:
            raise HTTPException(status_code=400, detail="Já existe um aluguel para este imóvel/proprietário neste período")
        
        novo_aluguel = AluguelSimples(
            ano=ano,
            mes=mes,
            imovel_id=imovel_id,
            proprietario_id=proprietario_id,
            taxa_administracao_total=taxa_administracao_total if 'taxa_administracao_total' in locals() else 0.0,
            valor_liquido_proprietario=valor_liquido if 'valor_liquido' in locals() else 0.0
            # taxa_administracao_proprietario será calculado automáticamente por trigger
        )
        
        db.add(novo_aluguel)
        db.commit()
        db.refresh(novo_aluguel)
        
        return {"sucesso": True, "mensagem": "Aluguel criado com sucesso", "id": novo_aluguel.id}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao criar aluguel: {str(e)}")

@router.get("/anos-disponiveis/")
async def obter_anos_disponiveis(db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Obter lista de anos que têm dados de aluguéis"""
    try:
        anos = db.query(AluguelSimples.ano).distinct().order_by(desc(AluguelSimples.ano)).all()
        anos_lista = [ano[0] for ano in anos if ano[0] is not None]
        print(f"📅 Anos disponíveis em dados: {anos_lista}")
        return {"success": True, "data": {'anos': anos_lista, 'total': len(anos_lista)}}
    except Exception as e:
        print(f"Erro em /alugueis/anos-disponiveis/: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro obtendo anos disponíveis: {str(e)}")

@router.get("/totais-por-imovel/")
async def obter_totais_por_imovel(
    ano: Optional[int] = Query(None, description="Filtrar por ano (por padrão último ano com dados)"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Filtrar por mês (por padrão último mês com dados)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Obter totais de aluguéis por imóvel para o último mês ou período especificado"""
    try:
        # Se não se especifica ano/mês, obter o último período disponível
        if not ano or not mes:
            ultimo_periodo = db.query(
                AluguelSimples.ano, 
                AluguelSimples.mes
            ).order_by(
                desc(AluguelSimples.ano), 
                desc(AluguelSimples.mes)
            ).first()
            
            if not ultimo_periodo:
                return []
            
            if not ano:
                ano = ultimo_periodo.ano
            if not mes:
                mes = ultimo_periodo.mes
        
        # Obter totais agrupados por imóvel para o período especificado
        resultado = db.query(
            AluguelSimples.imovel_id,
            func.sum(AluguelSimples.valor_liquido_proprietario).label('total_valor'),
            func.count(AluguelSimples.id).label('quantidade_proprietarios')
        ).filter(
            AluguelSimples.ano == ano,
            AluguelSimples.mes == mes
        ).group_by(
            AluguelSimples.imovel_id
        ).order_by(
            desc('total_valor')
        ).all()
        
        # Formatar resposta
        totais = []
        for row in resultado:
            imovel = db.query(Imovel).filter(Imovel.id == row.imovel_id).first()
            totais.append({
                'imovel_id': row.imovel_id,
                'nome_imovel': imovel.nome if imovel else None,
                'total_valor': float(row.total_valor),
                'quantidade_proprietarios': int(row.quantidade_proprietarios),
                'ano': ano,
                'mes': mes
            })
        
        return {"success": True, "data": {
            'periodo': {'ano': ano, 'mes': mes},
            'totais': totais,
            'total_imoveis': len(totais)
        }}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao obter totais por imóvel: {str(e)}")

@router.get("/totais-por-mes/")
async def obter_totais_por_mes(
    limite_meses: Optional[int] = Query(12, ge=1, le=24, description="Número de meses a incluir (máximo 24)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Obter totais de aluguéis agrupados por mês para o gráfico de tendências"""
    try:
        # Obter todos os períodos disponíveis ordenados por data
        resultado = db.query(
            AluguelSimples.ano,
            AluguelSimples.mes,
            func.sum(AluguelSimples.valor_liquido_proprietario).label('total_mes'),
            func.count(AluguelSimples.id).label('quantidade_alugueis')
        ).group_by(
            AluguelSimples.ano,
            AluguelSimples.mes
        ).order_by(
            desc(AluguelSimples.ano),
            desc(AluguelSimples.mes)
        ).limit(limite_meses).all()
        
        if not resultado:
            return {"success": True, "data": {
                'totais_mensais': [],
                'total_periodos': 0
            }}
        
        # Formatar resposta e inverter ordem para mostrar cronologicamente
        totais_mensais = []
        for row in reversed(resultado):
            periodo_label = formatar_periodo_label(row.ano, row.mes)
            
            totais_mensais.append({
                'ano': row.ano,
                'mes': row.mes,
                'periodo': periodo_label,
                'total_valor': float(row.total_mes),
                'quantidade_alugueis': int(row.quantidade_alugueis)
            })
        
        return {"success": True, "data": {
            'totais_mensais': totais_mensais,
            'total_periodos': len(totais_mensais)
        }}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao obter totais por mês: {str(e)}")

@router.get("/distribuicao-matriz/")
async def obter_distribuicao_matriz(
    ano: Optional[int] = Query(None, description="Filtrar por ano (por padrão último ano com dados)"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Filtrar por mês (por padrão último mês com dados)"),
    proprietario_id: Optional[int] = Query(None, description="Filtrar por ID de proprietário específico"),
    agregacao: Optional[str] = Query("mes_especifico", description="Tipo de agregação: 'mes_especifico', 'ano_completo', 'completo'"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Obter distribuição de aluguéis em formato matriz (proprietários vs imóveis) com agregação otimizada."""
    try:
        # Lógica de determinação de período
        ano_filtro, mes_filtro, periodo_texto = determinar_periodo_filtro(agregacao, ano, mes, db)

        # Query base para agregar os dados
        query = db.query(
            AluguelSimples.proprietario_id,
            Proprietario.nome.label('nome_proprietario'),
            AluguelSimples.imovel_id,
            Imovel.nome.label('nome_imovel'),
            func.sum(AluguelSimples.valor_liquido_proprietario).label('valor_total')
        ).join(Proprietario, AluguelSimples.proprietario_id == Proprietario.id)\
         .join(Imovel, AluguelSimples.imovel_id == Imovel.id)

        # Aplicar filtros
        if ano_filtro:
            query = query.filter(AluguelSimples.ano == ano_filtro)
        if mes_filtro:
            query = query.filter(AluguelSimples.mes == mes_filtro)
        if proprietario_id:
            query = query.filter(AluguelSimples.proprietario_id == proprietario_id)

        # Agrupar e ordenar
        query = query.group_by(
            AluguelSimples.proprietario_id,
            Proprietario.nome,
            AluguelSimples.imovel_id,
            Imovel.nome
        ).order_by(Proprietario.nome, Imovel.nome)

        aggregated_data = query.all()

        if not aggregated_data:
            return {"success": True, "data": {
                'periodo': {'ano': ano_filtro, 'mes': mes_filtro, 'tipo_agregacao': agregacao, 'descricao': periodo_texto},
                'proprietarios': [], 'imoveis': [], 'matriz': []
            }}

        # Processar dados agregados para formato de matriz
        proprietarios_map = {p.proprietario_id: {'id': p.proprietario_id, 'nome': p.nome_proprietario} for p in aggregated_data}
        imoveis_map = {i.imovel_id: {'id': i.imovel_id, 'nome': i.nome_imovel} for i in aggregated_data}
        
        matriz_map = {}
        for row in aggregated_data:
            if row.proprietario_id not in matriz_map:
                matriz_map[row.proprietario_id] = {
                    'proprietario_id': row.proprietario_id,
                    'nome_proprietario': row.nome_proprietario,
                    'valores': {im.nome: 0 for im in imoveis_map.values()},
                    'total': 0
                }
            matriz_map[row.proprietario_id]['valores'][row.nome_imovel] = float(row.valor_total or 0)
            matriz_map[row.proprietario_id]['total'] += float(row.valor_total or 0)

        # Ordenar e finalizar
        proprietarios_list = sorted(list(proprietarios_map.values()), key=lambda p: p['nome'])
        imoveis_list = sorted(list(imoveis_map.values()), key=lambda i: i['nome'])
        matriz_final = sorted(list(matriz_map.values()), key=lambda m: m['nome_proprietario'])

        return {"success": True, "data": {
            'periodo': {'ano': ano_filtro, 'mes': mes_filtro, 'tipo_agregacao': agregacao, 'descricao': periodo_texto},
            'proprietarios': proprietarios_list,
            'imoveis': imoveis_list,
            'matriz': matriz_final,
            'total_proprietarios': len(proprietarios_list),
            'total_imoveis': len(imoveis_list)
        }}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao obter distribuição matriz: {str(e)}")

def determinar_periodo_filtro(agregacao: str, ano: Optional[int], mes: Optional[int], db: Session):
    """Função auxiliar para determinar o período de filtro baseado na agregação."""
    if agregacao == "completo":
        return None, None, "Todos os períodos"
    elif agregacao == "ano_completo" and ano:
        return ano, None, f"Ano {ano}"
    else: # Mês específico
        if not ano or not mes:
            ultimo_periodo = db.query(AluguelSimples.ano, AluguelSimples.mes)\
                               .order_by(desc(AluguelSimples.ano), desc(AluguelSimples.mes))\
                               .first()
            if ultimo_periodo:
                ano = ano or ultimo_periodo.ano
                mes = mes or ultimo_periodo.mes
            else: # Nenhum dado na tabela
                return datetime.now().year, datetime.now().month, "Período atual (sem dados)"
        return ano, mes, f"{mes:02d}/{ano}"

@router.get("/aluguel/{aluguel_id}")
async def obter_aluguel_por_id(aluguel_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Obter um aluguel específico por ID"""
    aluguel = db.query(AluguelSimples).filter(AluguelSimples.id == aluguel_id).first()
    if not aluguel:
        raise HTTPException(status_code=404, detail="Aluguel não encontrado")
    
    return aluguel.to_dict()

@router.post("/")
async def criar_aluguel_dict(aluguel_data: dict, db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Criar um novo registro de aluguel"""
    try:
        # Criar objeto diretamente
        novo_aluguel = AluguelSimples(**aluguel_data)
        
        db.add(novo_aluguel)
        db.commit()
        db.refresh(novo_aluguel)
        
        return {
            "mensagem": "Aluguel criado com sucesso",
            "aluguel": novo_aluguel.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao criar aluguel: {str(e)}")

@router.put("/{aluguel_id}")
async def atualizar_aluguel(aluguel_id: int, aluguel_data: dict, db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Atualizar um aluguel existente"""
    try:
        aluguel = db.query(AluguelSimples).filter(AluguelSimples.id == aluguel_id).first()
        if not aluguel:
            raise HTTPException(status_code=404, detail="Aluguel não encontrado")
        
        # Atualizar campos
        for campo, valor in aluguel_data.items():
            if hasattr(aluguel, campo):
                setattr(aluguel, campo, valor)
        
        db.commit()
        db.refresh(aluguel)
        
        return {
            "mensagem": "Aluguel atualizado com sucesso",
            "aluguel": aluguel.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar aluguel: {str(e)}")

@router.delete("/{aluguel_id}")
async def excluir_aluguel(aluguel_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Excluir um aluguel"""
    try:
        aluguel = db.query(AluguelSimples).filter(AluguelSimples.id == aluguel_id).first()
        if not aluguel:
            raise HTTPException(status_code=404, detail="Aluguel não encontrado")
        
        db.delete(aluguel)
        db.commit()
        
        return {"mensagem": "Aluguel excluído com sucesso"}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao excluir aluguel: {str(e)}")

@router.post("/recalcular-taxas/")
async def recalcular_todas_as_taxas(db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Recalcula todas as taxas de administração por proprietário aplicando corretamente as participações"""
    try:
        resultado = CalculoService.recalcular_todas_as_taxas(db)
        
        return {
            "mensagem": "Recálculo de taxas completado",
            "resumo": resultado,
            "erros": resultado.get("erros")
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao recalcular taxas: {str(e)}")

@router.get("/ultimo-periodo/")
async def obter_ultimo_periodo(db: Session = Depends(get_db), current_user: Usuario = Depends(verify_token)):
    """Obter o último ano e mês disponível na base de dados"""
    try:
        ultimo_periodo = db.query(
            AluguelSimples.ano, 
            AluguelSimples.mes
        ).order_by(
            desc(AluguelSimples.ano), 
            desc(AluguelSimples.mes)
        ).first()
        
        if not ultimo_periodo:
            return {"success": True, "data": {"ano": None, "mes": None}}
        
        return {
            "success": True, 
            "data": {
                "ano": ultimo_periodo.ano,
                "mes": ultimo_periodo.mes
            }
        }
    except Exception as e:
        print(f"Erro em /alugueis/ultimo-periodo/: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro obtendo último período: {str(e)}")

@router.get("/distribuicao-todos-meses/")
async def obter_distribuicao_todos_meses(
    ano: int = Query(..., description="Ano para obter soma de todos os meses"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(verify_token)
):
    """Obter distribuição matriz de aluguéis com soma de todos os meses do ano especificado"""
    try:
        print(f"🔍 Buscando distribuição de todos os meses para ano {ano}")
        
        # Obter todos os registros do ano especificado
        alugueis = db.query(AluguelSimples).filter(
            AluguelSimples.ano == ano
        ).all()
        
        if not alugueis:
            return {"success": True, "data": {"matriz": [], "proprietarios": [], "imoveis": []}}
        
        # Agrupar por proprietário e imóvel, sumando todos los meses
        distribuicao = {}
        proprietarios_set = set()
        imoveis_set = set()
        
        for aluguel in alugueis:
            prop_id = aluguel.proprietario_id
            imovel_id = aluguel.imovel_id
            valor = aluguel.valor_liquido_proprietario or 0
            
            proprietarios_set.add(prop_id)
            imoveis_set.add(imovel_id)
            
            if prop_id not in distribuicao:
                distribuicao[prop_id] = {}
            if imovel_id not in distribuicao[prop_id]:
                distribuicao[prop_id][imovel_id] = 0
            
            distribuicao[prop_id][imovel_id] += valor
        
        # Converter a formato matriz
        proprietarios = []
        for prop_id in proprietarios_set:
            prop = db.query(Proprietario).filter(Proprietario.id == prop_id).first()
            if prop:
                proprietarios.append({
                    "proprietario_id": prop_id,
                    "nome": prop.nome
                })
        proprietarios.sort(key=lambda x: x['nome'])
        
        imoveis = []
        for imovel_id in imoveis_set:
            imovel = db.query(Imovel).filter(Imovel.id == imovel_id).first()
            if imovel:
                imoveis.append({
                    "id": imovel_id,
                    "nome": imovel.nome
                })
        imoveis.sort(key=lambda x: x['nome'])
        
        # Crear matriz
        matriz = []
        for prop in proprietarios:
            prop_id = prop["proprietario_id"]
            valores = {}
            for imovel in imoveis:
                imovel_id = imovel["id"]
                valores[imovel["nome"]] = distribuicao.get(prop_id, {}).get(imovel_id, 0)
            
            matriz.append({
                "proprietario_id": prop_id,
                "nome": prop["nome"],
                "valores": valores
            })
        
        print(f"✅ Matriz criada: {len(matriz)} proprietários, {len(imoveis)} imóveis")
        return {
            "success": True,
            "data": {
                "matriz": matriz,
                "proprietarios": proprietarios,
                "imoveis": imoveis
            }
        }
        
    except Exception as e:
        print(f"Erro em /alugueis/distribuicao-todos-meses/: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro obtendo distribuição de todos os meses: {str(e)}")